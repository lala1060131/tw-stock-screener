"""
台股低基期選股機器人 v3.0
功能：
  1. 動態抓取台股科技/電子/半導體/電機類前500大交易量股票
  2. 三大篩選門（財務防護 / 低基期(60日跌30%) / 轉機訊號(MA60+量能+營收)）
  3. 月營收 YoY 分析（FinMind API）
  4. 回測模組（歷史訊號勝率）
  5. LINE Notify 推播
  6. APScheduler 排程自動執行
"""

import yfinance as yf
import pandas as pd
import numpy as np
import json, os, time, requests, logging, sys
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import warnings
warnings.filterwarnings('ignore')

# 載入台股中文名稱對照表
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
if _THIS_DIR not in sys.path:
    sys.path.insert(0, _THIS_DIR)
try:
    from tw_stocks_zh import enrich_with_zh
except ImportError:
    def enrich_with_zh(result):
        result.setdefault("name_zh", "")
        result.setdefault("sector_zh", "")
        result.setdefault("subsector_zh", "")
        result.setdefault("display_name", result.get("name") or result.get("symbol", ""))
        return result

# ──────────────────────────────────────────────────────
# 路徑設定
# ──────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR    = os.path.join(BASE_DIR, 'data')
CONFIG_FILE = os.path.join(DATA_DIR, 'config.json')
RESULT_FILE = os.path.join(DATA_DIR, 'screening_result.json')
BACKTEST_FILE = os.path.join(DATA_DIR, 'backtest_result.json')
LOG_DIR     = os.path.join(BASE_DIR, 'logs')
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(LOG_DIR,  exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, 'screener.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────
# 設定檔管理
# ──────────────────────────────────────────────────────
DEFAULT_CONFIG = {
    "line_token": "",
    "finmind_token": "",
    "schedule_enabled": False,
    "schedule_time": "18:00",
    "schedule_days": ["Mon","Tue","Wed","Thu","Fri"],
    "max_workers": 6,
    "stock_pool_size": 500,
    "filters": {
        "debt_ratio_max": 50,
        "require_dividend": True,
        "require_eps_positive": True,
        "drop_60d_min": 30,
        "revenue_consec_months": 3
    },
    "last_run": None,
    "notify_min_gates": 2
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            # merge defaults
            for k, v in DEFAULT_CONFIG.items():
                if k not in cfg:
                    cfg[k] = v
            if "filters" not in cfg:
                cfg["filters"] = DEFAULT_CONFIG["filters"].copy()
            else:
                for fk, fv in DEFAULT_CONFIG["filters"].items():
                    if fk not in cfg["filters"]:
                        cfg["filters"][fk] = fv
            return cfg
        except:
            pass
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

# ──────────────────────────────────────────────────────
# 工具函數
# ──────────────────────────────────────────────────────
def safe_float(val, default=None):
    try:
        v = float(val)
        return None if (np.isnan(v) or np.isinf(v)) else v
    except:
        return default

# ──────────────────────────────────────────────────────
# 動態抓取台股科技/電子/半導體/電機類股票清單
# 以交易量排序，取前500支
# ──────────────────────────────────────────────────────

# 內建科技股基礎清單（作為 fallback）
try:
    from tw_hot_tech_zh import TW_HOT_TECH_MAP
    TECH_STOCKS_BASE = list(TW_HOT_TECH_MAP.keys())
except ImportError:
    TECH_STOCKS_BASE = [
        "2330","2303","2454","3661","3443","3711","6488","3037",
        "2382","3231","6669","2317","3017","3324","2308","2345",
    ]

# 擴展科技/電子/電機類股票清單
# 台灣證交所產業分類中，電子相關的類別代碼
TWSE_TECH_SECTORS = [
    "24",  # 電子工業 (含半導體、電腦、光電、通信、電子零件、資訊服務等)
    "31",  # 半導體業
    "32",  # 電腦及週邊設備業
    "33",  # 光電業
    "34",  # 通信網路業
    "35",  # 電子零組件業
    "36",  # 電子通路業
    "37",  # 資訊服務業
    "38",  # 其他電子業
    "15",  # 電機機械
]

def fetch_twse_tech_stocks():
    """從 TWSE 抓取電子/科技/電機類上市股票"""
    stocks = []
    try:
        # 方法1：TWSE 每日收盤行情（含成交量）
        url = "https://www.twse.com.tw/exchangeReport/STOCK_DAY_ALL?response=json"
        r = requests.get(url, timeout=15, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        if r.status_code == 200:
            data = r.json()
            if 'data' in data:
                for row in data['data']:
                    symbol = row[0].strip()
                    if symbol.isdigit() and len(symbol) == 4:
                        # 成交量在第2欄（去掉逗號）
                        try:
                            vol = int(row[2].replace(',', ''))
                        except:
                            vol = 0
                        stocks.append({'symbol': symbol, 'volume': vol})
    except Exception as e:
        log.warning(f"TWSE 全市場資料抓取失敗: {e}")

    if not stocks:
        # 方法2：fallback 用 TWSE 個股日成交資訊
        try:
            url = "https://www.twse.com.tw/exchangeReport/MI_INDEX?response=json&type=ALLBUT0999"
            r = requests.get(url, timeout=15, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            if r.status_code == 200:
                data = r.json()
                tables = data.get('tables', [])
                for table in tables:
                    if 'data' in table:
                        for row in table['data']:
                            symbol = row[0].strip().replace('"', '')
                            if symbol.isdigit() and len(symbol) == 4:
                                try:
                                    vol = int(str(row[2]).replace(',', ''))
                                except:
                                    vol = 0
                                stocks.append({'symbol': symbol, 'volume': vol})
        except Exception as e:
            log.warning(f"TWSE MI_INDEX 抓取失敗: {e}")

    return stocks

def fetch_tpex_tech_stocks():
    """從 TPEx 抓取上櫃電子/科技類股票"""
    stocks = []
    try:
        today = datetime.now()
        date_str = f"{today.year - 1911}/{today.month:02d}/{today.day:02d}"
        url = f"https://www.tpex.org.tw/web/stock/aftertrading/daily_close_quotes/stk_quote_result.php?l=zh-tw&d={date_str}&_=1"
        r = requests.get(url, timeout=15, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        if r.status_code == 200:
            data = r.json()
            aa_data = data.get('aaData', [])
            for row in aa_data:
                symbol = str(row[0]).strip()
                if symbol.isdigit() and len(symbol) == 4:
                    try:
                        vol = int(str(row[8]).replace(',', ''))
                    except:
                        vol = 0
                    stocks.append({'symbol': symbol, 'volume': vol})
    except Exception as e:
        log.warning(f"TPEx 資料抓取失敗: {e}")
    return stocks

# 科技/電子/電機相關股票代碼範圍判斷
# 台股上市：2xxx (電子/電機), 3xxx (電子), 4xxx (部分電子), 5xxx (部分), 6xxx (電子)
# 上櫃：3xxx, 4xxx, 5xxx, 6xxx, 8xxx
TECH_SYMBOL_PREFIXES_LISTED = ['23','24','25','26','27','28','29','30','31','32','33','34','35','36','37','38','39']

def is_likely_tech_stock(symbol):
    """判斷股票代碼是否可能屬於科技/電子/電機類"""
    # 已知在我們的科技股基礎清單中
    if symbol in TECH_STOCKS_BASE:
        return True
    # 台股電子類通常是 2xxx(部分), 3xxx, 4xxx(部分), 5xxx(部分), 6xxx
    first_digit = symbol[0] if symbol else ''
    # 2300~2999 多為電子/電機
    if symbol.startswith('2') and int(symbol) >= 2300:
        return True
    # 3000~3999 幾乎都是電子
    if first_digit == '3':
        return True
    # 4000~4999 部分電子（生技也在此）
    if first_digit == '4' and int(symbol) >= 4900:
        return True
    # 5000~5999 部分電子
    if symbol.startswith('5') and int(symbol) >= 5200 and int(symbol) <= 5600:
        return True
    # 6000~6999 多為電子/科技
    if first_digit == '6':
        return True
    # 8000~8999 上櫃電子
    if first_digit == '8':
        return True
    return False

def get_tw_stock_list(size=500):
    """
    取得要篩選的科技/電子/電機類股票清單。
    優先以交易量排序，取前 size 支。
    如果動態抓取失敗，使用內建清單。
    """
    log.info(f"取得台股科技類清單（目標 {size} 支，以交易量排序）...")

    all_stocks = []

    # 抓取上市
    twse = fetch_twse_tech_stocks()
    if twse:
        tech_twse = [s for s in twse if is_likely_tech_stock(s['symbol'])]
        all_stocks.extend(tech_twse)
        log.info(f"  上市科技股: {len(tech_twse)} 支（全市場 {len(twse)} 支）")

    # 抓取上櫃
    tpex = fetch_tpex_tech_stocks()
    if tpex:
        tech_tpex = [s for s in tpex if is_likely_tech_stock(s['symbol'])]
        all_stocks.extend(tech_tpex)
        log.info(f"  上櫃科技股: {len(tech_tpex)} 支（全市場 {len(tpex)} 支）")

    if all_stocks:
        # 依交易量排序（大到小）
        all_stocks.sort(key=lambda x: x['volume'], reverse=True)
        # 去重
        seen = set()
        result = []
        for s in all_stocks:
            if s['symbol'] not in seen:
                seen.add(s['symbol'])
                result.append(s['symbol'])
        # 確保基礎清單中的熱門股都在前面
        final = []
        for s in TECH_STOCKS_BASE:
            if s in seen:
                final.append(s)
                seen.discard(s)
        # 補上剩餘的（按交易量排序）
        for s in result:
            if s not in final:
                final.append(s)
            if len(final) >= size:
                break
        log.info(f"  最終清單: {len(final[:size])} 支")
        return final[:size]
    else:
        # Fallback：使用內建清單
        log.warning("  動態抓取失敗，使用內建科技股清單")
        return TECH_STOCKS_BASE[:size]

# ──────────────────────────────────────────────────────
# 第一門：財務防護門
# ──────────────────────────────────────────────────────
def check_financial_guard(ticker_obj, info, cfg_filters):
    result = {
        "pass_financial": False,
        "eps_5yr_positive": False,
        "eps_years_positive": 0,
        "debt_ratio": None,
        "debt_ratio_ok": False,
        "has_dividend": False,
        "dividend_yield": None,
        "payout_ratio": None,
    }
    try:
        # ── 負債比 ──────────────────────────────────────
        debt_max = cfg_filters.get("debt_ratio_max", 50)
        total_debt   = safe_float(info.get('totalDebt'))
        total_assets = safe_float(info.get('totalAssets'))
        if total_debt is not None and total_assets and total_assets > 0:
            dr = total_debt / total_assets * 100
            result["debt_ratio"]    = round(dr, 1)
            result["debt_ratio_ok"] = bool(dr < debt_max)
        else:
            de = safe_float(info.get('debtToEquity'))
            if de is not None and de >= 0:
                de_n = de / 100
                dr = de_n / (1 + de_n) * 100
                result["debt_ratio"]    = round(dr, 1)
                result["debt_ratio_ok"] = bool(dr < debt_max)
            else:
                result["debt_ratio"]    = None
                result["debt_ratio_ok"] = True  # 缺失寬鬆

        # ── 股利（含殖利率合理性檢查） ────────────────────
        if cfg_filters.get("require_dividend", True):
            div_yield = safe_float(info.get('dividendYield'))
            if div_yield is not None and div_yield > 0:
                # yfinance 回傳的 dividendYield 應為小數（如 0.05 = 5%）
                # 但某些台股會回傳已經是百分比的值（如 5.0 = 5%）
                # 合理性檢查：如果值 > 1，代表已經是百分比格式
                if div_yield > 1:
                    # 已經是百分比，直接使用
                    yield_pct = div_yield
                else:
                    # 小數格式，轉為百分比
                    yield_pct = div_yield * 100

                # 殖利率合理性檢查：超過30%視為異常資料
                if yield_pct > 30:
                    # 可能是資料異常，嘗試用股價和股利重新計算
                    annual_div = safe_float(info.get('dividendRate'))
                    current_price = safe_float(info.get('currentPrice') or info.get('regularMarketPrice'))
                    if annual_div and current_price and current_price > 0:
                        yield_pct = annual_div / current_price * 100
                        if yield_pct > 30:
                            yield_pct = None  # 仍然異常，放棄
                    else:
                        yield_pct = None  # 無法驗證，放棄

                if yield_pct is not None and yield_pct > 0:
                    result["has_dividend"]   = True
                    result["dividend_yield"] = round(yield_pct, 2)
                else:
                    # 最後嘗試用 dividendRate
                    annual_div = safe_float(info.get('dividendRate'))
                    current_price = safe_float(info.get('currentPrice') or info.get('regularMarketPrice'))
                    if annual_div and annual_div > 0 and current_price and current_price > 0:
                        yield_pct = annual_div / current_price * 100
                        if 0 < yield_pct <= 30:
                            result["has_dividend"]   = True
                            result["dividend_yield"] = round(yield_pct, 2)
            else:
                # dividendYield 為 0 或 None，嘗試用 dividendRate 計算
                annual_div = safe_float(info.get('dividendRate'))
                current_price = safe_float(info.get('currentPrice') or info.get('regularMarketPrice'))
                if annual_div and annual_div > 0 and current_price and current_price > 0:
                    yield_pct = annual_div / current_price * 100
                    if 0 < yield_pct <= 30:
                        result["has_dividend"]   = True
                        result["dividend_yield"] = round(yield_pct, 2)

            payout = safe_float(info.get('payoutRatio'))
            if payout is not None and payout > 0:
                payout_pct = payout * 100 if payout < 5 else payout  # 合理性
                if payout_pct <= 200:  # 超過200%視為異常
                    result["payout_ratio"] = round(payout_pct, 1)
        else:
            result["has_dividend"] = True  # 不要求配息

        # ── EPS 連年正 ────────────────────────────────────
        if cfg_filters.get("require_eps_positive", True):
            years_pos = 0
            try:
                fin = ticker_obj.financials
                if fin is not None and not fin.empty:
                    ni_row = next((r for r in fin.index
                                   if 'Net Income' in str(r) or 'net_income' in str(r).lower()), None)
                    if ni_row is not None:
                        vals = fin.loc[ni_row].values
                        years_pos = sum(1 for v in vals[:5] if safe_float(v, -1) > 0)
                    else:
                        eps = safe_float(info.get('trailingEps'))
                        years_pos = 1 if (eps and eps > 0) else 0
                else:
                    eps = safe_float(info.get('trailingEps'))
                    years_pos = 1 if (eps and eps > 0) else 0
            except:
                eps = safe_float(info.get('trailingEps'))
                years_pos = 1 if (eps and eps > 0) else 0

            result["eps_years_positive"] = years_pos
            result["eps_5yr_positive"]   = bool(years_pos >= 3)
        else:
            result["eps_5yr_positive"] = True

        result["pass_financial"] = bool(
            result["eps_5yr_positive"] and
            result["debt_ratio_ok"] and
            result["has_dividend"]
        )
    except Exception as e:
        log.debug(f"財務門例外: {e}")
    return result

# ──────────────────────────────────────────────────────
# 第二門：低基期門（60日跌幅 > 30%）
# ──────────────────────────────────────────────────────
def check_low_base(hist_2y, cfg_filters):
    """
    計算 60 個交易日跌幅。drop_60d_pct 為負值代表下跌。
    pass_low_base = 跌幅超過門檻（預設 -30%）
    """
    result = {
        "pass_low_base":   False,
        "drop_60d_pct":    None,
        "drawdown_ok":     False,
        "price_60d_ago":   None,
        "current_price":   None,
    }
    try:
        if hist_2y is None or len(hist_2y) < 61:
            return result
        close = hist_2y['Close']
        cur = float(close.iloc[-1])
        past = float(close.iloc[-61])  # 60 個交易日前
        if past > 0:
            drop = (cur - past) / past * 100
            result["drop_60d_pct"]  = round(drop, 1)
            result["price_60d_ago"] = round(past, 2)
            result["current_price"] = round(cur, 2)
            drop_min = cfg_filters.get("drop_60d_min", 30)  # 正數門檻
            result["drawdown_ok"]   = bool(drop <= -abs(drop_min))
            result["pass_low_base"] = result["drawdown_ok"]
    except Exception as e:
        log.debug(f"低基期門例外: {e}")
    return result

# ──────────────────────────────────────────────────────
# 第三門：轉機訊號門（站上MA60 + 量能上升 + 營收成長）
# ──────────────────────────────────────────────────────
def check_turnaround(hist_2y, info, revenue_data):
    """
    轉機訊號門：
    - 站上季線 MA60
    - 近月量能上升
    - 營收有成長跡象（YoY>0 或 連續雙增）
    三者中至少滿足「站上MA60 + (量能上升 或 營收成長)」即通過
    """
    result = {
        "pass_turnaround":   False,
        "above_60ma":        False,
        "ma60":              None,
        "volume_change_pct": None,
        "volume_rising":     False,
        "rsi14":             None,
        "golden_cross":      False,
        "vol_3m_avg":        None,
        "ma5yr_price":       None,
        "above_5yr_ma":      False,
    }
    try:
        if hist_2y is None or len(hist_2y) < 61:
            return result

        close  = hist_2y['Close']
        volume = hist_2y['Volume']

        # MA60
        ma60 = float(close.rolling(60).mean().iloc[-1])
        cur  = float(close.iloc[-1])
        if not np.isnan(ma60):
            result["ma60"]       = round(ma60, 2)
            result["above_60ma"] = bool(cur > ma60)

        # 黃金交叉（MA20 > MA60 且前一天 MA20 < MA60）
        if len(close) >= 61:
            ma20 = close.rolling(20).mean()
            ma60s = close.rolling(60).mean()
            if len(ma20) >= 2 and not np.isnan(float(ma20.iloc[-1])):
                cur_cross  = float(ma20.iloc[-1]) > float(ma60s.iloc[-1])
                prev_cross = float(ma20.iloc[-2]) <= float(ma60s.iloc[-2])
                result["golden_cross"] = bool(cur_cross and prev_cross)

        # 5年均線
        if len(close) >= 1260:
            ma5y = float(close.rolling(1260).mean().iloc[-1])
            result["ma5yr_price"]  = round(ma5y, 2)
            result["above_5yr_ma"] = bool(cur > ma5y)
        elif len(close) >= 120:
            ma5y = float(close.mean())
            result["ma5yr_price"]  = round(ma5y, 2)
            result["above_5yr_ma"] = bool(cur > ma5y)

        # 量能趨勢
        if len(volume) >= 60:
            v_recent = float(volume.iloc[-30:].mean())
            v_prev   = float(volume.iloc[-60:-30].mean())
            v_3m     = float(volume.iloc[-63:].mean())
            if v_prev > 0:
                chg = (v_recent - v_prev) / v_prev * 100
                result["volume_change_pct"] = round(chg, 1)
                result["volume_rising"]     = bool(chg > 0)
            result["vol_3m_avg"] = round(v_3m / 1e4, 1)  # 萬股

        # RSI-14
        if len(close) >= 15:
            delta  = close.diff()
            gain   = delta.clip(lower=0).rolling(14).mean()
            loss   = (-delta.clip(upper=0)).rolling(14).mean()
            rs     = gain / loss.replace(0, np.nan)
            rsi    = 100 - 100 / (1 + rs)
            rsi_v  = safe_float(rsi.iloc[-1])
            result["rsi14"] = round(rsi_v, 1) if rsi_v else None

        # 轉機訊號通過條件：站上MA60 + (量能上升 或 營收有成長)
        revenue_growing = bool(
            revenue_data.get("revenue_yoy") is not None and revenue_data.get("revenue_yoy", -999) > 0
        ) or revenue_data.get("double_growth_3m", False) or revenue_data.get("revenue_turning", False)

        result["pass_turnaround"] = bool(
            result["above_60ma"] and (result["volume_rising"] or revenue_growing)
        )
    except Exception as e:
        log.debug(f"轉機門例外: {e}")
    return result

# ──────────────────────────────────────────────────────
# 月營收 YoY（FinMind API / fallback）
# ──────────────────────────────────────────────────────
def get_revenue_yoy(symbol, finmind_token=""):
    """
    透過 FinMind API 取得月營收年增率
    無 token 時使用免費額度（有限制）
    """
    result = {
        "revenue_yoy":       None,
        "revenue_mom":       None,
        "revenue_3m_trend":  None,  # 近3月平均YoY
        "revenue_turning":   False, # YoY由負轉正
        "revenue_data_ok":   False,
        "consec_double_months": 0,  # 連續「MoM>0 且 YoY>0」月數
        "double_growth_3m":  False, # 是否連3月雙增
    }
    try:
        headers = {}
        params = {
            "dataset": "TaiwanStockMonthRevenue",
            "data_id": symbol,
            "start_date": (datetime.now() - timedelta(days=400)).strftime("%Y-%m-%d"),
            "end_date":   datetime.now().strftime("%Y-%m-%d"),
        }
        if finmind_token:
            params["token"] = finmind_token

        r = requests.get(
            "https://api.finmindtrade.com/api/v4/data",
            params=params, headers=headers, timeout=10
        )
        if r.status_code != 200:
            return result

        data = r.json().get('data', [])
        if not data or len(data) < 2:
            return result

        df = pd.DataFrame(data)
        df['date'] = pd.to_datetime(df['date'])
        df = df.sort_values('date')
        df['revenue'] = pd.to_numeric(df['revenue'], errors='coerce')

        if len(df) >= 13:
            latest    = float(df['revenue'].iloc[-1])
            prev_year = float(df['revenue'].iloc[-13])
            prev_mon  = float(df['revenue'].iloc[-2])
            if prev_year > 0:
                yoy = (latest - prev_year) / prev_year * 100
                result["revenue_yoy"] = round(yoy, 1)
            if prev_mon > 0:
                mom = (latest - prev_mon) / prev_mon * 100
                result["revenue_mom"] = round(mom, 1)

            # 近3月平均YoY
            if len(df) >= 15:
                yoys = []
                for i in range(-1, -4, -1):
                    cur = float(df['revenue'].iloc[i])
                    yr_ago = float(df['revenue'].iloc[i-12])
                    if yr_ago > 0:
                        yoys.append((cur - yr_ago) / yr_ago * 100)
                if yoys:
                    result["revenue_3m_trend"] = round(np.mean(yoys), 1)

            # 由負轉正判斷：上月YoY<0，本月YoY>0
            if len(df) >= 14 and result["revenue_yoy"] is not None:
                prev2 = float(df['revenue'].iloc[-2])
                yr_ago2 = float(df['revenue'].iloc[-14])
                if yr_ago2 > 0:
                    prev_yoy = (prev2 - yr_ago2) / yr_ago2 * 100
                    result["revenue_turning"] = bool(prev_yoy < 0 and result["revenue_yoy"] > 0)

            # 連續「MoM>0 且 YoY>0」月數（從最新月份往回算）
            consec = 0
            for i in range(-1, -7, -1):  # 最多檢查近 6 個月
                if abs(i) >= len(df) or abs(i - 1) >= len(df) or abs(i - 12) >= len(df):
                    break
                cur_rev = float(df['revenue'].iloc[i])
                prev_m  = float(df['revenue'].iloc[i - 1])    # 上個月
                yr_ago  = float(df['revenue'].iloc[i - 12])   # 去年同月
                if prev_m > 0 and yr_ago > 0:
                    mom_i = (cur_rev - prev_m) / prev_m * 100
                    yoy_i = (cur_rev - yr_ago) / yr_ago * 100
                    if mom_i > 0 and yoy_i > 0:
                        consec += 1
                    else:
                        break
                else:
                    break
            result["consec_double_months"] = consec
            result["double_growth_3m"] = bool(consec >= 3)

            result["revenue_data_ok"] = True
    except Exception as e:
        log.debug(f"FinMind {symbol} 例外: {e}")
    return result

# ──────────────────────────────────────────────────────
# 回測模組
# ──────────────────────────────────────────────────────
def backtest_signal(symbol, hold_days=60):
    """
    回測：當訊號觸發時（站上60MA），持有 hold_days 後的報酬率
    """
    try:
        t = yf.Ticker(f"{symbol}.TW")
        hist = t.history(period="5y", interval="1d")
        if hist is None or len(hist) < 120:
            return None

        close = hist['Close']
        ma60  = close.rolling(60).mean()

        signals = []
        for i in range(60, len(close) - hold_days - 1):
            # 訊號：從60MA下方穿越到上方
            prev_below = float(close.iloc[i-1]) < float(ma60.iloc[i-1])
            cur_above  = float(close.iloc[i])   > float(ma60.iloc[i])
            if prev_below and cur_above:
                entry = float(close.iloc[i])
                exit_ = float(close.iloc[i + hold_days])
                ret   = (exit_ - entry) / entry * 100
                signals.append({
                    "date":   hist.index[i].strftime("%Y-%m-%d"),
                    "entry":  round(entry, 2),
                    "exit":   round(exit_, 2),
                    "return": round(ret, 2),
                    "win":    ret > 0,
                })

        if not signals:
            return None

        rets   = [s["return"] for s in signals]
        wins   = sum(1 for s in signals if s["win"])
        return {
            "symbol":       symbol,
            "total_signals": len(signals),
            "win_rate":      round(wins / len(signals) * 100, 1),
            "avg_return":    round(np.mean(rets), 2),
            "max_return":    round(max(rets), 2),
            "min_return":    round(min(rets), 2),
            "std_return":    round(np.std(rets), 2),
            "signals":       signals[-10:],  # 最近10筆
        }
    except Exception as e:
        log.debug(f"回測 {symbol} 例外: {e}")
        return None

def run_backtest(symbols=None, hold_days=60, max_workers=4):
    """批次回測"""
    if symbols is None:
        # 從現有結果取通過2門以上的股票
        if os.path.exists(RESULT_FILE):
            with open(RESULT_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            symbols = [r['symbol'] for r in data.get('results', [])
                       if r.get('gates_passed', 0) >= 2][:30]
        else:
            symbols = TECH_STOCKS_BASE[:20]

    log.info(f"開始回測 {len(symbols)} 支股票（持有{hold_days}天）...")
    results = []

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(backtest_signal, s, hold_days): s for s in symbols}
        for i, fut in enumerate(as_completed(futures)):
            s = futures[fut]
            try:
                r = fut.result(timeout=60)
                if r:
                    results.append(r)
                    log.info(f"  [{i+1}/{len(symbols)}] {s} 勝率={r['win_rate']}% 均報酬={r['avg_return']}%")
                else:
                    log.info(f"  [{i+1}/{len(symbols)}] {s} 無訊號")
            except Exception as e:
                log.warning(f"  [{i+1}/{len(symbols)}] {s} 失敗: {e}")

    results.sort(key=lambda x: -x['win_rate'])

    summary = {
        "run_at":       datetime.now().isoformat(),
        "hold_days":    hold_days,
        "total_tested": len(symbols),
        "with_signals": len(results),
        "overall_win_rate": round(np.mean([r['win_rate'] for r in results]), 1) if results else 0,
        "overall_avg_return": round(np.mean([r['avg_return'] for r in results]), 2) if results else 0,
        "results": results,
    }
    with open(BACKTEST_FILE, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2, default=str)
    log.info(f"✅ 回測完成！平均勝率={summary['overall_win_rate']}%，平均報酬={summary['overall_avg_return']}%")
    return summary

# ──────────────────────────────────────────────────────
# LINE Notify 推播
# ──────────────────────────────────────────────────────
def send_line_notify(token, message):
    if not token:
        log.warning("LINE token 未設定，跳過推播")
        return False
    try:
        r = requests.post(
            "https://notify-api.line.me/api/notify",
            headers={"Authorization": f"Bearer {token}"},
            data={"message": message},
            timeout=10
        )
        ok = r.status_code == 200
        log.info(f"LINE 推播 {'成功' if ok else '失敗'} (status={r.status_code})")
        return ok
    except Exception as e:
        log.error(f"LINE 推播例外: {e}")
        return False

def build_line_message(summary, min_gates=2):
    results = [r for r in summary.get('results', []) if r.get('gates_passed', 0) >= min_gates]
    now = datetime.now().strftime("%Y/%m/%d %H:%M")
    lines = [
        f"\n🤖 台股低基期選股機器人 v3",
        f"📅 {now}",
        f"──────────────────",
        f"篩選總數: {summary.get('total_screened', 0)} 支",
        f"⭐ 全部通過(3/3): {summary.get('all_pass_count', 0)} 支",
        f"👁 通過2門: {summary.get('two_gates_count', 0)} 支",
        f"──────────────────",
    ]
    if results:
        lines.append(f"📋 推薦標的（{min_gates}門以上）：")
        for r in results[:10]:
            gates = "⭐" * r['gates_passed'] + f" {r['gates_passed']}/3"
            drop = f"60日跌{r.get('drop_60d_pct','–')}%" if r.get('drop_60d_pct') is not None else ""
            rev = f"營收YoY={r.get('revenue_yoy','–')}%" if r.get('revenue_yoy') is not None else ""
            dy = f"殖利率={r.get('dividend_yield','–')}%" if r.get('dividend_yield') is not None else ""
            name = r.get('display_name') or r.get('name','')
            lines.append(f"  {r['symbol']} {name[:8]}  {gates}")
            detail = "  ".join([x for x in [drop, rev, dy] if x])
            if detail:
                lines.append(f"    {detail}")
    else:
        lines.append("今日無符合條件標的")
    lines.append("──────────────────")
    lines.append("⚠ 僅供參考，非投資建議")
    return "\n".join(lines)

# ──────────────────────────────────────────────────────
# 核心篩選流程（單支）
# ──────────────────────────────────────────────────────
def screen_single(symbol, cfg=None):
    if cfg is None:
        cfg = load_config()
    filters = cfg.get("filters", DEFAULT_CONFIG["filters"])
    finmind_token = cfg.get("finmind_token", "")

    try:
        t    = yf.Ticker(f"{symbol}.TW")
        info = t.info

        name       = info.get('longName') or info.get('shortName') or symbol
        market_cap = safe_float(info.get('marketCap'))
        sector     = info.get('sector', '未分類')
        industry   = info.get('industry', '未知產業')
        pe_ratio   = safe_float(info.get('trailingPE') or info.get('forwardPE'))

        # 抓取2年歷史資料
        hist_2y = None
        try:
            hist_2y = t.history(period="2y", interval="1d")
        except:
            pass

        # 三門篩選
        # 門1：財務防護門
        f = check_financial_guard(t, info, filters)

        # 門2：低基期門（60日跌幅>30%）
        lb = check_low_base(hist_2y, filters)

        # 月營收（供門3使用）
        rev = get_revenue_yoy(symbol, finmind_token)

        # 門3：轉機訊號門（站上MA60 + 量能/營收）
        ta = check_turnaround(hist_2y, info, rev)

        # 計算通過門數
        gates_passed = sum([f["pass_financial"], lb["pass_low_base"], ta["pass_turnaround"]])
        all_pass = bool(gates_passed == 3)

        # 綜合評分（0~100）
        score = 0
        score += 25 if f["pass_financial"] else 0
        score += 30 if lb["pass_low_base"] else 0
        score += 25 if ta["pass_turnaround"] else 0
        # 營收強度加分（近3月平均YoY，最多10分）
        r3m = rev.get("revenue_3m_trend")
        if r3m is not None and r3m > 0:
            score += min(int(r3m / 3), 10)
        # 殖利率加分（最多10分）
        dy = f.get("dividend_yield")
        if dy is not None and dy > 0:
            score += min(int(dy), 10)

        # 取得當前股價（優先從低基期門取得）
        current_price = lb.get("current_price")
        if current_price is None and hist_2y is not None and len(hist_2y) > 0:
            current_price = round(float(hist_2y['Close'].iloc[-1]), 2)

        result = {
            "symbol":        symbol,
            "name":          name,
            "sector":        sector,
            "industry":      industry,
            "market_cap":    market_cap,
            "market_cap_b":  round(market_cap / 1e8, 1) if market_cap else None,
            "pe_ratio":      round(pe_ratio, 1) if pe_ratio else None,
            "current_price": current_price,
            # 財務防護門
            "pass_financial":    f["pass_financial"],
            "eps_5yr_positive":  f["eps_5yr_positive"],
            "eps_years_positive":f["eps_years_positive"],
            "debt_ratio":        f["debt_ratio"],
            "debt_ratio_ok":     f["debt_ratio_ok"],
            "has_dividend":      f["has_dividend"],
            "dividend_yield":    f["dividend_yield"],
            "payout_ratio":      f["payout_ratio"],
            # 低基期門
            "pass_low_base":     lb["pass_low_base"],
            "drop_60d_pct":      lb["drop_60d_pct"],
            "drawdown_ok":       lb["drawdown_ok"],
            "price_60d_ago":     lb["price_60d_ago"],
            # 轉機訊號門
            "pass_turnaround":   ta["pass_turnaround"],
            "above_60ma":        ta["above_60ma"],
            "ma60":              ta["ma60"],
            "golden_cross":      ta["golden_cross"],
            "volume_change_pct": ta["volume_change_pct"],
            "volume_rising":     ta["volume_rising"],
            "rsi14":             ta["rsi14"],
            "vol_3m_avg":        ta["vol_3m_avg"],
            # 月營收
            "revenue_yoy":       rev.get("revenue_yoy"),
            "revenue_mom":       rev.get("revenue_mom"),
            "revenue_3m_trend":  rev.get("revenue_3m_trend"),
            "revenue_turning":   rev.get("revenue_turning"),
            "consec_double_months": rev.get("consec_double_months"),
            "double_growth_3m":  rev.get("double_growth_3m"),
            "revenue_data_ok":   rev.get("revenue_data_ok"),
            # 綜合
            "gates_passed": gates_passed,
            "all_pass":     all_pass,
            "score":        min(score, 100),
        }
        return enrich_with_zh(result)
    except Exception as e:
        log.warning(f"screen_single {symbol} 失敗: {e}")
        return enrich_with_zh({
            "symbol": symbol, "name": symbol,
            "gates_passed": 0, "all_pass": False,
            "score": 0, "error": str(e)
        })

# ──────────────────────────────────────────────────────
# 批次篩選主流程
# ──────────────────────────────────────────────────────
def run_screening(symbols=None, cfg=None):
    if cfg is None:
        cfg = load_config()
    if symbols is None:
        symbols = get_tw_stock_list(cfg.get("stock_pool_size", 500))

    max_workers = cfg.get("max_workers", 6)
    log.info(f"=== 台股低基期選股機器人 v3 啟動 ===")
    log.info(f"篩選清單: {len(symbols)} 支  | workers: {max_workers}")

    results = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(screen_single, s, cfg): s for s in symbols}
        for i, fut in enumerate(as_completed(futures)):
            s = futures[fut]
            try:
                r = fut.result(timeout=90)
                results.append(r)
                mark = "⭐" if r.get("all_pass") else f"  {r.get('gates_passed',0)}/3"
                log.info(f"  [{i+1:>3}/{len(symbols)}] {s:>6} {mark}  {r.get('name','')[:20]}")
            except Exception as e:
                results.append({"symbol": s, "name": s, "gates_passed": 0,
                                 "all_pass": False, "score": 0, "error": str(e)})

    # 排序：評分最高優先，其次通過門數、跌幅深
    results.sort(key=lambda x: (
        -x.get("score", 0),
        -x.get("gates_passed", 0),
        (x.get("drop_60d_pct") if x.get("drop_60d_pct") is not None else 999),
    ))

    summary = {
        "run_at":           datetime.now().isoformat(),
        "total_screened":   len(symbols),
        "total_results":    len(results),
        "all_pass_count":   sum(1 for r in results if r.get("all_pass")),
        "two_gates_count":  sum(1 for r in results if r.get("gates_passed") == 2),
        "one_gate_count":   sum(1 for r in results if r.get("gates_passed") == 1),
        "results":          results,
    }
    with open(RESULT_FILE, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2, default=str)

    log.info(f"✅ 篩選完成！全通過:{summary['all_pass_count']} | 2門:{summary['two_gates_count']} | 1門:{summary['one_gate_count']}")

    # LINE 推播
    line_token = cfg.get("line_token", "")
    min_gates  = cfg.get("notify_min_gates", 2)
    if line_token:
        msg = build_line_message(summary, min_gates)
        send_line_notify(line_token, msg)

    # 更新 config 最後執行時間
    cfg["last_run"] = datetime.now().isoformat()
    save_config(cfg)

    # 順便產 Excel
    try:
        export_excel(summary)
    except Exception as e:
        log.warning(f"Excel 匯出失敗: {e}")

    return summary

# ──────────────────────────────────────────────────────
# Excel 匯出
# ──────────────────────────────────────────────────────
def export_excel(summary=None):
    if summary is None:
        if not os.path.exists(RESULT_FILE):
            return None
        with open(RESULT_FILE, 'r', encoding='utf-8') as f:
            summary = json.load(f)

    results = [r for r in summary.get("results", []) if r.get("gates_passed", 0) > 0]
    if not results:
        return None

    rows = []
    for r in results:
        rows.append({
            "股票代碼": r.get("symbol",""),
            "公司名稱": r.get("display_name") or r.get("name",""),
            "產業別":   r.get("sector_zh") or r.get("sector",""),
            "市值(億)": r.get("market_cap_b"),
            "綜合評分": r.get("score"),
            "通過門數": r.get("gates_passed",0),
            "完全通過": "✅" if r.get("all_pass") else "",
            # 財務防護門
            "EPS連年正": "✅" if r.get("eps_5yr_positive") else "❌",
            "獲利年數": r.get("eps_years_positive"),
            "負債比(%)": r.get("debt_ratio"),
            "有配息":   "✅" if r.get("has_dividend") else "❌",
            "現金殖利率(%)": r.get("dividend_yield"),
            # 低基期門
            "當前股價": r.get("current_price"),
            "60日前股價": r.get("price_60d_ago"),
            "60日跌幅(%)": r.get("drop_60d_pct"),
            "跌幅達標(>30%)": "✅" if r.get("drawdown_ok") else "❌",
            # 轉機訊號門
            "站上MA60": "✅" if r.get("above_60ma") else "❌",
            "量能上升": "✅" if r.get("volume_rising") else "❌",
            "量能變化(%)": r.get("volume_change_pct"),
            "RSI14": r.get("rsi14"),
            # 月營收
            "月營收YoY(%)": r.get("revenue_yoy"),
            "月營收MoM(%)": r.get("revenue_mom"),
            "3月均YoY(%)":  r.get("revenue_3m_trend"),
            "連續雙增月數":  r.get("consec_double_months"),
            "連3月雙增":    "✅" if r.get("double_growth_3m") else "❌",
        })

    df = pd.DataFrame(rows)
    fname = f"低基期選股_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    out   = os.path.join(DATA_DIR, fname)

    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='全部結果', index=False)
        df[df["完全通過"]=="✅"].to_excel(writer, sheet_name='完全通過3門', index=False)
        df[df["通過門數"]==2].to_excel(writer, sheet_name='觀察名單2門', index=False)

        from openpyxl.styles import PatternFill, Font, Alignment
        header = PatternFill("solid", fgColor="1A237E")
        for sn in writer.sheets:
            ws = writer.sheets[sn]
            for cell in ws[1]:
                cell.fill = header
                cell.font = Font(color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                mx = max((len(str(c.value or "")) for c in col), default=4)
                ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 22)

    log.info(f"📊 Excel 匯出: {out}")
    return fname

# ──────────────────────────────────────────────────────
# APScheduler 排程
# ──────────────────────────────────────────────────────
_scheduler = None

def start_scheduler():
    global _scheduler
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger

    cfg = load_config()
    if not cfg.get("schedule_enabled"):
        log.info("排程未啟用")
        return None

    if _scheduler and _scheduler.running:
        _scheduler.shutdown(wait=False)

    _scheduler = BackgroundScheduler(timezone="Asia/Taipei")
    t = cfg.get("schedule_time", "18:00").split(":")
    hour, minute = int(t[0]), int(t[1])

    days_map = {
        "Mon":"mon","Tue":"tue","Wed":"wed",
        "Thu":"thu","Fri":"fri","Sat":"sat","Sun":"sun"
    }
    days = cfg.get("schedule_days", ["Mon","Tue","Wed","Thu","Fri"])
    day_of_week = ",".join(days_map.get(d, d.lower()) for d in days)

    _scheduler.add_job(
        run_screening,
        CronTrigger(day_of_week=day_of_week, hour=hour, minute=minute),
        id="daily_screen",
        replace_existing=True,
    )
    _scheduler.start()
    log.info(f"✅ 排程啟動：每週 {day_of_week} {hour:02d}:{minute:02d}")
    return _scheduler

def stop_scheduler():
    global _scheduler
    if _scheduler and _scheduler.running:
        _scheduler.shutdown(wait=False)
        log.info("排程已停止")

def get_next_run_time():
    global _scheduler
    if not _scheduler or not _scheduler.running:
        return None
    jobs = _scheduler.get_jobs()
    if not jobs:
        return None
    return jobs[0].next_run_time.isoformat() if jobs[0].next_run_time else None

# ──────────────────────────────────────────────────────
# CLI 入口
# ──────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    mode = sys.argv[1] if len(sys.argv) > 1 else "full"

    if mode == "quick":
        run_screening(TECH_STOCKS_BASE[:15], load_config())
    elif mode == "backtest":
        run_backtest(hold_days=int(sys.argv[2]) if len(sys.argv) > 2 else 60)
    elif mode == "line_test":
        cfg = load_config()
        send_line_notify(cfg.get("line_token",""), "🤖 LINE Notify 測試訊息 — 台股選股機器人連線成功！")
    elif mode == "excel":
        export_excel()
    elif mode == "list":
        stocks = get_tw_stock_list(500)
        print(f"取得 {len(stocks)} 支股票：{stocks[:10]}...")
    else:
        run_screening()
